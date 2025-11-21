"""
Reports Utility Functions
"""
import csv
import time
from io import BytesIO, StringIO
from django.core.exceptions import ValidationError
from django.utils import timezone
from django.db import connection
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment


def execute_report_query(query, parameters=None):
    """
    Execute SQL query for report data
    
    Args:
        query: SQL query string
        parameters: Query parameters
    
    Returns:
        list: Query results
    """
    with connection.cursor() as cursor:
        if parameters:
            cursor.execute(query, parameters)
        else:
            cursor.execute(query)
        
        columns = [col[0] for col in cursor.description]
        results = []
        for row in cursor.fetchall():
            results.append(dict(zip(columns, row)))
        
        return results


def generate_pdf_report(title, data, columns, filename=None):
    """
    Generate PDF report
    
    Args:
        title: Report title
        data: Report data (list of dicts)
        columns: List of column definitions
        filename: Output filename
    
    Returns:
        BytesIO: PDF file buffer
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Add title
    title_paragraph = Paragraph(title, styles['Title'])
    elements.append(title_paragraph)
    elements.append(Spacer(1, 0.2 * inch))
    
    # Prepare table data
    table_data = []
    headers = [col.get('label', col['key']) for col in columns]
    table_data.append(headers)
    
    for row in data:
        row_data = [str(row.get(col['key'], '')) for col in columns]
        table_data.append(row_data)
    
    # Create table
    if table_data:
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    return buffer


def generate_excel_report(title, data, columns, filename=None):
    """
    Generate Excel report
    
    Args:
        title: Report title
        data: Report data (list of dicts)
        columns: List of column definitions
        filename: Output filename
    
    Returns:
        BytesIO: Excel file buffer
    """
    buffer = BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = title[:31]  # Excel sheet name limit
    
    # Add title
    worksheet['A1'] = title
    worksheet['A1'].font = Font(size=16, bold=True)
    worksheet.merge_cells('A1:' + chr(64 + len(columns)) + '1')
    
    # Add headers
    for idx, col in enumerate(columns, start=1):
        cell = worksheet.cell(row=3, column=idx)
        cell.value = col.get('label', col['key'])
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for row_idx, row in enumerate(data, start=4):
        for col_idx, col in enumerate(columns, start=1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.value = row.get(col['key'], '')
    
    # Auto-adjust column widths
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value or '')) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
    
    workbook.save(buffer)
    buffer.seek(0)
    
    return buffer


def generate_csv_report(data, columns):
    """
    Generate CSV report
    
    Args:
        data: Report data (list of dicts)
        columns: List of column definitions
    
    Returns:
        str: CSV data
    """
    output = StringIO()
    writer = csv.writer(output)
    
    # Write headers
    headers = [col.get('label', col['key']) for col in columns]
    writer.writerow(headers)
    
    # Write data
    for row in data:
        row_data = [row.get(col['key'], '') for col in columns]
        writer.writerow(row_data)
    
    return output.getvalue()


def generate_html_report(title, data, columns):
    """
    Generate HTML report
    
    Args:
        title: Report title
        data: Report data (list of dicts)
        columns: List of column definitions
    
    Returns:
        str: HTML content
    """
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>{title}</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; }}
            h1 {{ color: #333; }}
            table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
            th {{ background-color: #4CAF50; color: white; }}
            tr:nth-child(even) {{ background-color: #f2f2f2; }}
        </style>
    </head>
    <body>
        <h1>{title}</h1>
        <table>
            <thead>
                <tr>
    """
    
    for col in columns:
        html += f"<th>{col.get('label', col['key'])}</th>"
    
    html += """
                </tr>
            </thead>
            <tbody>
    """
    
    for row in data:
        html += "<tr>"
        for col in columns:
            html += f"<td>{row.get(col['key'], '')}</td>"
        html += "</tr>"
    
    html += """
            </tbody>
        </table>
    </body>
    </html>
    """
    
    return html


def get_client_ip(request):
    """Get client IP address from request"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def log_report_access(user, template, action, request, generated_report=None, metadata=None):
    """
    Log report access
    
    Args:
        user: User accessing the report
        template: Report template
        action: Action performed
        request: HTTP request object
        generated_report: Generated report (optional)
        metadata: Additional metadata (optional)
    """
    from .models import ReportAccessLog
    
    ReportAccessLog.objects.create(
        user=user,
        template=template,
        generated_report=generated_report,
        action=action,
        ip_address=get_client_ip(request),
        user_agent=request.META.get('HTTP_USER_AGENT', ''),
        metadata=metadata or {}
    )


def calculate_next_run(schedule_type, scheduled_time, timezone_str='UTC'):
    """
    Calculate next run time for scheduled report
    
    Args:
        schedule_type: Type of schedule (daily, weekly, monthly, etc.)
        scheduled_time: Time to run
        timezone_str: Timezone string
    
    Returns:
        datetime: Next run time
    """
    import pytz
    from datetime import datetime, timedelta
    
    tz = pytz.timezone(timezone_str)
    now = timezone.now().astimezone(tz)
    
    # Combine date and time
    next_run = tz.localize(datetime.combine(now.date(), scheduled_time))
    
    # If time has passed today, schedule for tomorrow
    if next_run <= now:
        if schedule_type == 'daily':
            next_run += timedelta(days=1)
        elif schedule_type == 'weekly':
            next_run += timedelta(weeks=1)
        elif schedule_type == 'monthly':
            # Add one month
            month = next_run.month
            year = next_run.year
            if month == 12:
                month = 1
                year += 1
            else:
                month += 1
            next_run = next_run.replace(year=year, month=month)
    
    return next_run.astimezone(pytz.UTC)


def validate_report_parameters(parameters, parameter_definitions):
    """
    Validate report parameters against definitions
    
    Args:
        parameters: Parameters to validate
        parameter_definitions: Parameter definitions
    
    Returns:
        dict: Validated parameters
    
    Raises:
        ValidationError: If validation fails
    """
    validated = {}
    
    for param_name, param_def in parameter_definitions.items():
        value = parameters.get(param_name)
        required = param_def.get('required', False)
        param_type = param_def.get('type', 'text')
        
        if required and value is None:
            raise ValidationError(f"Parameter '{param_name}' is required")
        
        if value is not None:
            # Type validation
            if param_type == 'number':
                try:
                    validated[param_name] = int(value)
                except (ValueError, TypeError):
                    raise ValidationError(f"Parameter '{param_name}' must be a number")
            elif param_type == 'date':
                from django.utils.dateparse import parse_date
                date_value = parse_date(str(value))
                if not date_value:
                    raise ValidationError(f"Parameter '{param_name}' must be a valid date")
                validated[param_name] = date_value.isoformat()
            else:
                validated[param_name] = value
        else:
            validated[param_name] = param_def.get('default')
    
    return validated
