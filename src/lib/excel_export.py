"""Excel Export Utilities for EMIS"""
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.models.reports import QuarterlyReport
from src.models.billing import Bill
from src.lib.logging import get_logger

logger = get_logger(__name__)


class ExcelExporter:
    """Utility class for exporting data to Excel"""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=12)
        self.title_font = Font(bold=True, size=14)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_quarterly_report(
        self,
        report: QuarterlyReport,
        filepath: str
    ) -> str:
        """Export quarterly report to Excel"""
        
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Title
        ws_summary['A1'] = f"Quarterly Financial Report - Q{report.quarter} {report.financial_year}"
        ws_summary['A1'].font = self.title_font
        ws_summary.merge_cells('A1:D1')
        
        ws_summary['A2'] = f"Period: {report.start_date.strftime('%d %b %Y')} to {report.end_date.strftime('%d %b %Y')}"
        ws_summary.merge_cells('A2:D2')
        
        # Executive Summary
        row = 4
        ws_summary[f'A{row}'] = "Executive Summary"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        
        row += 1
        summary_data = [
            ['Metric', 'Value'],
            ['Total Income', f"₹{report.total_income:,.2f}"],
            ['Total Expenses', f"₹{report.total_expenses:,.2f}"],
            ['Net Profit/Loss', f"₹{report.net_profit_loss:,.2f}"],
            ['Students', report.student_count],
            ['Employees', report.employee_count],
            ['Avg Fee per Student', f"₹{report.average_fee_per_student:,.2f}"],
            ['Fee Collection Rate', f"{report.fee_collection_rate:.1f}%"],
        ]
        
        for data_row in summary_data:
            ws_summary[f'A{row}'] = data_row[0]
            ws_summary[f'B{row}'] = data_row[1]
            
            if row == 5:  # Header
                ws_summary[f'A{row}'].fill = self.header_fill
                ws_summary[f'B{row}'].fill = self.header_fill
                ws_summary[f'A{row}'].font = self.header_font
                ws_summary[f'B{row}'].font = self.header_font
            
            ws_summary[f'A{row}'].border = self.border
            ws_summary[f'B{row}'].border = self.border
            
            row += 1
        
        # Auto-size columns
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20
        
        # Income sheet
        ws_income = wb.create_sheet("Income Breakdown")
        
        ws_income['A1'] = "Income Breakdown"
        ws_income['A1'].font = self.title_font
        ws_income.merge_cells('A1:B1')
        
        income_data = [
            ['Category', 'Amount'],
            ['Tuition Fees', f"₹{report.tuition_fee_income:,.2f}"],
            ['Admission Fees', f"₹{report.admission_fee_income:,.2f}"],
            ['Exam Fees', f"₹{report.exam_fee_income:,.2f}"],
            ['Library Fees', f"₹{report.library_fee_income:,.2f}"],
            ['Other Fees', f"₹{report.other_fee_income:,.2f}"],
            ['Miscellaneous', f"₹{report.miscellaneous_income:,.2f}"],
            ['Total', f"₹{report.total_income:,.2f}"],
        ]
        
        row = 3
        for data_row in income_data:
            ws_income[f'A{row}'] = data_row[0]
            ws_income[f'B{row}'] = data_row[1]
            
            if row == 3:  # Header
                ws_income[f'A{row}'].fill = self.header_fill
                ws_income[f'B{row}'].fill = self.header_fill
                ws_income[f'A{row}'].font = self.header_font
                ws_income[f'B{row}'].font = self.header_font
            elif row == len(income_data) + 2:  # Total row
                ws_income[f'A{row}'].font = Font(bold=True)
                ws_income[f'B{row}'].font = Font(bold=True)
            
            ws_income[f'A{row}'].border = self.border
            ws_income[f'B{row}'].border = self.border
            
            row += 1
        
        ws_income.column_dimensions['A'].width = 30
        ws_income.column_dimensions['B'].width = 20
        
        # Expense sheet
        ws_expense = wb.create_sheet("Expense Breakdown")
        
        ws_expense['A1'] = "Expense Breakdown"
        ws_expense['A1'].font = self.title_font
        ws_expense.merge_cells('A1:B1')
        
        expense_data = [
            ['Category', 'Amount'],
            ['Salaries', f"₹{report.salary_expenses:,.2f}"],
            ['Maintenance', f"₹{report.maintenance_expenses:,.2f}"],
            ['Utilities', f"₹{report.utility_expenses:,.2f}"],
            ['Infrastructure', f"₹{report.infrastructure_expenses:,.2f}"],
            ['Administrative', f"₹{report.administrative_expenses:,.2f}"],
            ['Emergency', f"₹{report.emergency_expenses:,.2f}"],
            ['Other', f"₹{report.other_expenses:,.2f}"],
            ['Total', f"₹{report.total_expenses:,.2f}"],
        ]
        
        row = 3
        for data_row in expense_data:
            ws_expense[f'A{row}'] = data_row[0]
            ws_expense[f'B{row}'] = data_row[1]
            
            if row == 3:  # Header
                ws_expense[f'A{row}'].fill = self.header_fill
                ws_expense[f'B{row}'].fill = self.header_fill
                ws_expense[f'A{row}'].font = self.header_font
                ws_expense[f'B{row}'].font = self.header_font
            elif row == len(expense_data) + 2:  # Total row
                ws_expense[f'A{row}'].font = Font(bold=True)
                ws_expense[f'B{row}'].font = Font(bold=True)
            
            ws_expense[f'A{row}'].border = self.border
            ws_expense[f'B{row}'].border = self.border
            
            row += 1
        
        ws_expense.column_dimensions['A'].width = 30
        ws_expense.column_dimensions['B'].width = 20
        
        # Detailed data sheet (if available)
        if report.income_details or report.expense_details:
            ws_details = wb.create_sheet("Detailed Data")
            
            ws_details['A1'] = "Detailed Income and Expense Data"
            ws_details['A1'].font = self.title_font
            ws_details.merge_cells('A1:C1')
            
            row = 3
            ws_details['A3'] = "Income Details"
            ws_details['A3'].font = Font(bold=True)
            row = 4
            
            if report.income_details:
                ws_details[f'A{row}'] = "Category"
                ws_details[f'B{row}'] = "Amount"
                ws_details[f'A{row}'].fill = self.header_fill
                ws_details[f'B{row}'].fill = self.header_fill
                ws_details[f'A{row}'].font = self.header_font
                ws_details[f'B{row}'].font = self.header_font
                row += 1
                
                for category, amount in report.income_details.items():
                    ws_details[f'A{row}'] = category
                    ws_details[f'B{row}'] = f"₹{amount:,.2f}"
                    row += 1
            
            row += 2
            ws_details[f'A{row}'] = "Expense Details"
            ws_details[f'A{row}'].font = Font(bold=True)
            row += 1
            
            if report.expense_details:
                ws_details[f'A{row}'] = "Category"
                ws_details[f'B{row}'] = "Amount"
                ws_details[f'A{row}'].fill = self.header_fill
                ws_details[f'B{row}'].fill = self.header_fill
                ws_details[f'A{row}'].font = self.header_font
                ws_details[f'B{row}'].font = self.header_font
                row += 1
                
                for category, amount in report.expense_details.items():
                    ws_details[f'A{row}'] = category
                    ws_details[f'B{row}'] = f"₹{amount:,.2f}"
                    row += 1
            
            ws_details.column_dimensions['A'].width = 40
            ws_details.column_dimensions['B'].width = 20
        
        # Save workbook
        wb.save(filepath)
        
        logger.info(f"Exported quarterly report to Excel: {filepath}")
        
        return filepath
    
    def export_bills(
        self,
        bills: List[Bill],
        filepath: str
    ) -> str:
        """Export bills list to Excel"""
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Bills"
        
        # Title
        ws['A1'] = "Bills Report"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:I1')
        
        ws['A2'] = f"Generated on: {datetime.now().strftime('%d %b %Y %I:%M %p')}"
        ws.merge_cells('A2:I2')
        
        # Headers
        headers = [
            'Bill Number', 'Date', 'Type', 'Student/Employee',
            'Subtotal', 'Tax', 'Total', 'Paid', 'Due', 'Status'
        ]
        
        row = 4
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        row = 5
        for bill in bills:
            entity = f"Student {bill.student_id}" if bill.student_id else f"Employee {bill.employee_id}"
            
            data = [
                bill.bill_number,
                bill.bill_date.strftime('%d-%b-%Y'),
                bill.bill_type.value.replace('_', ' ').title(),
                entity,
                bill.subtotal,
                bill.tax_amount,
                bill.total_amount,
                bill.amount_paid,
                bill.amount_due,
                bill.status.value.title()
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = self.border
                
                # Format currency columns
                if col in [5, 6, 7, 8, 9]:
                    cell.number_format = '₹#,##0.00'
            
            row += 1
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 15
        
        # Save workbook
        wb.save(filepath)
        
        logger.info(f"Exported {len(bills)} bills to Excel: {filepath}")
        
        return filepath
